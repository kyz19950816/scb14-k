# -*- coding: utf-8 -*-
# @Time ： 2020/7/15 14:02
# @Auth ： K
# @File ：lesson7.15-2.py
# @QQ ：378986003
# @weixin：kyz599520
# ************************

'''
接口自动化步骤
1、excel测试用例准备ok，代码自动读取测试数据     ----read_data()
2、发送接口请求，得到响应信息  ----api_fun()
3、断言：实际结果vs预期结果   ---通过/不通过
4、写入通过/不通过  ----excel
'''
import requests
import openpyxl

#封装成函数（读取测试用例函数）
def read_data(filename,sheetname):      #定义函数
    wb = openpyxl.load_workbook(filename)  # 加载工作薄   ---文档名字
    sheet = wb[sheetname]  # 获取表单
    max_row = sheet.max_row            # 获取最大行数
    case_list = []                       #创建空列表，存放测试用例
    for i in range (2,max_row+1):        #循环遍历
        dict1 = dict(
        case_id = sheet.cell(row=i, column=1).value,   # 获取case_id
        url = sheet.cell(row=i,column=5).value,       # 获取url
        data = sheet.cell(row=i,column=6).value,      # 获取data
        expect = sheet.cell(row=i,column=7).value,    # 获取expect
        )                                       #dict（）打包成整体变成字典格式
        case_list.append(dict1 )     #每循环一次，就把读到的字典数据存放到这个list里
    return case_list  # 返回测试用例列表




#执行接口函数
def api_fun(url,data):
    headers = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}  #字典格式
    res=requests.post(url=url ,json= data,headers=headers )#接收post方法
    request=res.json()#响应正文
    return request    #返回值




#封装函数（写入结果）
def write_result(filename,sheetname,row,column,final_result):   #定义参数， 根据上面读取列表函数获取
    wb = openpyxl.load_workbook(filename)   # 加载工作薄   ---文档名字
    sheet = wb[sheetname]                               # 获取表单
    sheet.cell(row, column).value = final_result          #找到对应位置写入结果
    wb.save('test_case_api.xlsx')                #保存， （需要先关闭文档）



#执行测试用例并回写实际结果

# cases = read_data('test_case_api.xlsx','register')    #调用读取测试用例，获取所以有测试用例数据保存到变量
#需要循环遍历提取case_id、url、data、expect
#---eval()   ---运行被字符串包裹的表达式
# for case in cases :                      #for循环取出case_id、url、data、expect
#     case_id =  case.get('case_id')     #第二种取id方法 ：case['case_id']
#     url = case.get('url')             #获取rul
#     data = eval(case.get('data') )          #---eval() 去掉被字符串包裹的表达式   #获取data
#     expect = eval(case.get('expect'))       # 获取预期结果  expect
#     expect_msg = expect.get('msg')    #获取预期结果中的msg信息
#     #print(type(data))                #查看data数据类型
#     real_result = api_fun(url=url,data =data )  #调用发送接口请求函数，返回结果用变量real_result接收
#     real_msg = real_result.get('msg')          #获取实际结果中的msg信息
#     # print(case_id ,url ,data ,expect )      #分别取出对应的值
#     # print(real_result)               #打印执行接口测试结果
#     print('预期结果中的msg:{}'.format(expect_msg ) )
#     print('实际结果中的msg:{}'.format(real_msg ))
#     if real_msg == expect_msg :                #if来对比预期结果和实际结果
#         print('第{}这条测试用例执行通过！'.format(case_id ) )
#         final_re='Passed'
#     else:
#         print('第{}这条测试用例执行不通过！'.format(case_id ) )
#         final_re='Failed'
#     write_result('test_case_api.xlsx','register',case_id+1,8,final_re)   #调用函数判断写入结果
#     print('*' * 25)  # 打印*号分隔


#封装函数（执行测试用例并回写实际结果）
def execule_fun(filename,sheetname):
    cases = read_data(filename,sheetname)  # 调用读取测试用例，获取所以有测试用例数据保存到变量

    # 需要循环遍历提取case_id、url、data、expect
    # ---eval()   ---运行被字符串包裹的表达式
    for case in cases:  # for循环取出case_id、url、data、expect
        case_id = case.get('case_id')  # 第二种取id方法 ：case['case_id']
        url = case.get('url')  # 获取rul
        data = eval(case.get('data'))  # ---eval() 去掉被字符串包裹的表达式   #获取data
        expect = eval(case.get('expect'))  # 获取预期结果  expect
        expect_msg = expect.get('msg')  # 获取预期结果中的msg信息
        # print(type(data))                #查看data数据类型
        real_result = api_fun(url=url, data=data)  # 调用发送接口请求函数，返回结果用变量real_result接收
        real_msg = real_result.get('msg')  # 获取实际结果中的msg信息
        # print(case_id ,url ,data ,expect )      #分别取出对应的值
        # print(real_result)               #打印执行接口测试结果
        print('预期结果中的msg:{}'.format(expect_msg))
        print('实际结果中的msg:{}'.format(real_msg))
        if real_msg == expect_msg:  # if来对比预期结果和实际结果
            print('第{}这条测试用例执行通过！'.format(case_id))
            final_re = 'Passed'
        else:
            print('第{}这条测试用例执行不通过！'.format(case_id))
            final_re = 'Failed'
        write_result(filename,sheetname, case_id + 1, 8, final_re)  # 调用函数判断写入结果
        print('*' * 25)  # 打印*号分隔


execule_fun('test_case_api.xlsx', 'login')
